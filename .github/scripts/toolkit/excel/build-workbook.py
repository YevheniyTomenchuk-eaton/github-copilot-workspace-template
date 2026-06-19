#!/usr/bin/env python3
"""Build a polished, on-brand Excel .xlsx workbook from a JSON content spec.

Usage:
    python build-workbook.py <spec.json> <output.xlsx>

Each sheet renders a colored banner title, a frozen bold header row, banded
data rows, auto-fit columns, an auto-filter, and an optional bold totals row.
Per-column `formats` apply number / currency / percent / date styling.

Spec shape (see the office-documents skill for the full reference):
    {
      "sheets": [
        {
          "name": "Pipeline",
          "title": "Q3 Sales Pipeline",
          "headers": ["Deal", "Owner", "Value", "Probability", "Close date"],
          "formats": ["text", "text", "currency", "percent", "date"],
          "rows": [["Acme", "Dana", 120000, 0.8, "2026-08-15"]],
          "totals": ["Total", "", 120000, "", ""]
        }
      ]
    }

Emits machine-readable KEY=value lines on success:
    OUTPUT=<absolute path to the .xlsx>
    SHEETS=<number of sheets>
    ROWS=<total data rows across all sheets>

Requires: openpyxl  (pip install openpyxl)
"""

import datetime
import json
import os
import sys

PRIMARY = "1F4E79"
ACCENT = "00B3A4"
INK = "1A2233"
BAND = "EEF3F9"
RULE = "D5DEEA"
WHITE = "FFFFFF"

FORMAT_CODES = {
    "currency": "$#,##0",
    "percent": "0%",
    "number": "#,##0",
    "date": "yyyy-mm-dd",
    "text": None,
}


def _coerce(value, fmt):
    if fmt == "date" and isinstance(value, str):
        try:
            return datetime.datetime.strptime(value, "%Y-%m-%d")
        except ValueError:
            return value
    return value


def build(spec_path, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR=openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    with open(spec_path, encoding="utf-8") as spec_file:
        spec = json.load(spec_file)

    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor=PRIMARY)
    banner_fill = PatternFill("solid", fgColor=INK)
    band_fill = PatternFill("solid", fgColor=BAND)
    header_font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    banner_font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    total_font = Font(bold=True, color=INK, size=11, name="Calibri")
    thin = Side(style="thin", color=RULE)
    cell_border = Border(bottom=thin)
    top_border = Border(top=Side(style="medium", color=ACCENT))
    center = Alignment(horizontal="left", vertical="center")

    sheets = spec.get("sheets", [])
    total_rows = 0

    for index, sheet_spec in enumerate(sheets):
        name = sheet_spec.get("name") or f"Sheet{index + 1}"
        worksheet = workbook.create_sheet(title=name[:31])
        worksheet.sheet_properties.tabColor = PRIMARY

        headers = sheet_spec.get("headers", [])
        formats = sheet_spec.get("formats", [])
        column_count = max(len(headers), 1)

        cursor = 1

        if sheet_spec.get("title"):
            worksheet.merge_cells(
                start_row=cursor, start_column=1,
                end_row=cursor, end_column=column_count,
            )
            banner = worksheet.cell(row=cursor, column=1, value=sheet_spec["title"])
            banner.fill = banner_fill
            banner.font = banner_font
            banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            worksheet.row_dimensions[cursor].height = 28
            cursor += 1

        header_row = None
        if headers:
            header_row = cursor
            for column, label in enumerate(headers, start=1):
                cell = worksheet.cell(row=cursor, column=column, value=label)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            worksheet.row_dimensions[cursor].height = 22
            cursor += 1

        for row_index, row in enumerate(sheet_spec.get("rows", [])):
            for column, value in enumerate(row, start=1):
                fmt = formats[column - 1] if column - 1 < len(formats) else "text"
                cell = worksheet.cell(row=cursor, column=column, value=_coerce(value, fmt))
                cell.border = cell_border
                code = FORMAT_CODES.get(fmt)
                if code:
                    cell.number_format = code
                if row_index % 2 == 1:
                    cell.fill = band_fill
            cursor += 1
            total_rows += 1

        totals = sheet_spec.get("totals")
        if totals:
            for column, value in enumerate(totals, start=1):
                fmt = formats[column - 1] if column - 1 < len(formats) else "text"
                cell = worksheet.cell(row=cursor, column=column, value=_coerce(value, fmt))
                cell.font = total_font
                cell.border = top_border
                code = FORMAT_CODES.get(fmt)
                if code:
                    cell.number_format = code
            cursor += 1

        for column in range(1, column_count + 1):
            letter = get_column_letter(column)
            longest = len(str(headers[column - 1])) if column - 1 < len(headers) else 8
            for row in sheet_spec.get("rows", []):
                if column - 1 < len(row):
                    longest = max(longest, len(str(row[column - 1])))
            worksheet.column_dimensions[letter].width = min(max(longest + 4, 12), 48)

        if header_row:
            worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)
            last_column = get_column_letter(column_count)
            worksheet.auto_filter.ref = f"A{header_row}:{last_column}{header_row}"

    if not workbook.sheetnames:
        workbook.create_sheet(title="Sheet1")

    output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    workbook.save(output_path)

    print(f"OUTPUT={output_path}")
    print(f"SHEETS={len(workbook.sheetnames)}")
    print(f"ROWS={total_rows}")


def main():
    if len(sys.argv) != 3:
        print("ERROR=usage: python build-workbook.py <spec.json> <output.xlsx>")
        sys.exit(1)
    build(sys.argv[1], sys.argv[2])


if __name__ == "__main__":
    main()
